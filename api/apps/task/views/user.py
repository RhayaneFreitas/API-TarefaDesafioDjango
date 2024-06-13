from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import JSONRenderer
from django.http import HttpResponse
import json
from django.utils import timezone
from datetime import (
    datetime,
    timedelta,
    date
)
from api.apps.task.serializers.user import (
    TasksSerializer,
    TaskResponsibleSerializer,
    TaskReportFilterSerializer
    )
from api.apps.task.serializers import user
from api.apps.task.models import (
    TaskProfile,
    User,
    
    )
from api.apps.task.models.task import TaskResponsible
from api.apps.task.models import task
from rest_framework.authentication import TokenAuthentication #Udemy
from rest_framework import generics
import django_filters.rest_framework # Filters
from django_filters.rest_framework import DjangoFilterBackend

# from api.apps.task.serializers import TasksSerializer
from rest_framework import (
    permissions,
    status,
    viewsets,
    filters,
    )

from django.db.models import (
    Count,
    Value,
    F,
    Q
    )

#Entendendo parte Agregate Django:
from django.db.models.aggregates import (
    Avg,
    Sum,
    Count,
    Min,
    Max
)
from django.db.models.functions import Concat
from django.shortcuts import get_object_or_404
from openpyxl import Workbook


class TaskView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, format=None):
        
        tasks = TaskProfile.objects.all()
        serializer = TasksSerializer(tasks, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = TasksSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        try:
            task = TaskProfile.objects.get(pk=pk)
        except TaskProfile.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
        serializer = TasksSerializer(task, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        try:
            task = TaskProfile.objects.get(pk=pk)
        except TaskProfile.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
        task.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ExportData(APIView):
    renderer_classes = [JSONRenderer]

    def get(self, request):
        tasks = TaskProfile.objects.all()
        formato = request.GET.get('formato')
        serializer = TasksSerializer(tasks,many=True)
        formatos = {
            "json": self.get_json,
            "txt": self.get_txt
        }
        try:
            content, content_type = formatos[formato](serializer)
            response = HttpResponse(content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="tasks.{formato}"'
            
            return response
        
        except KeyError:
            return Response({"detail": "Formato não suportado "},status=status.HTTP_400_BAD_REQUEST)

               
    def get_json(self, serializer):
        # Serializa os dados em JSON
        content = JSONRenderer().render(serializer.data)
        return content, "application/json"
    
    def get_txt(self, serializer):
        # Gera o conteúdo em formato de texto
        content = []
        for task in serializer.data:
            content.append(
                f'Titulo: {task["title"]}\n'
                f'Descricao: {task["description"]}\n'
                f'Prazo: {task["deadline"]}\n'
                f'Data de Lançamento: {task["release"]}\n'
                f'Concluida: {task["completed"]}\n'
                f'Finalizada Em: {task["finished_in"]}\n'
                f'Finalizada Por: {task["finished_by"]}\n'
                f'Criada Em: {task["created_in"]}\n'
                f'Atualizada: {task["updated"]}\n'
                f'Responsavel: {task["responsible"]}\n'
                '---\n'
            )
        content = ''.join(content)

        return content, 'text/plain; charset=utf-8'

class TasksCreatedFinishedByUserView(APIView):

    def get(self, request):
        format = request.GET.get('formato')
        
        if format == 'excel':
            return self.export_to_excel()
        elif format == 'json':
            return self.export_to_json()
        # Se os parâmetros de filtro não estiverem presentes, retornar todos os dados
        elif not request.GET.get('created_by') and not request.GET.get('finished_by'):
            return self.get_all_data(request)
        # Caso contrário, aplicar filtro
        else:
            return self.get_filter(request)

    def get_all_data(self, request):
        
        users = User.objects.all()
        
        output_data = []
        
        # Para cada usuário, conta as tarefas criadas e finalizadas
        for user in users:
            created_task_count = TaskProfile.objects.filter(created_by=user).count()
            finished_task_count = TaskProfile.objects.filter(finished_by=user).count()
            
            # Adiciona os dados do usuário à lista de saída
            output_data.append({
                'id': user.id,
                'name': user.name,
                'tasks_created': created_task_count,
                'tasks_finished': finished_task_count
            })
        
        return Response(output_data)

    def get_filter(self, request):
        created_by = request.GET.get('created_by')
        finished_by = request.GET.get('finished_by')

        # Inicializa os contadores
        created_task_count = 0
        finished_task_count = 0
        
        # Filtro para contar as tarefas criadas pelo usuário, se fornecido
        if created_by:
            created_task_count = TaskProfile.objects.filter(created_by_id=created_by).count()
        
        # Filtro para contar as tarefas finalizadas pelo usuário, se fornecido
        if finished_by:
            finished_task_count = TaskProfile.objects.filter(finished_by_id=finished_by).count()
        
        output_data = {
            "total_tasks_created_by_user": created_task_count,
            "total_tasks_finished_by_user": finished_task_count
        }

        return Response(output_data)
    
    def export_to_excel(self):
        users = User.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks Created and Finished by User Report"
        ws.append(["ID", "Name", "Tasks Created", "Tasks Finished"])

        for user in users:
            created_task_count = TaskProfile.objects.filter(created_by=user).count()
            finished_task_count = TaskProfile.objects.filter(finished_by=user).count()

            ws.append([user.id, user.name, created_task_count, finished_task_count])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=tasks_created_finished_by_user_report.xlsx'
        wb.save(response)

        return response

    def export_to_json(self):
        users = User.objects.all()
        output_data = []

        for user in users:
            created_task_count = TaskProfile.objects.filter(created_by=user).count()
            finished_task_count = TaskProfile.objects.filter(finished_by=user).count()

            output_data.append({
                'id': user.id,
                'name': user.name,
                'tasks_created': created_task_count,
                'tasks_finished': finished_task_count
            })

        return Response(output_data)
        
class ActivitiesByResponsibleView(APIView):

    def get(self, request):
        format = request.GET.get('formato')

        if format == 'excel':
            return self.export_to_excel()
        elif format == 'json':
            return self.export_to_json()
        else:
            return Response({"detail": "Formato não suportado."}, status=status.HTTP_400_BAD_REQUEST)

    def export_to_excel(self):
        # Quantidade de atividades por responsável
        responsible_tasks = User.objects.annotate(
            task_count=Count('tasks_responsible')
        ).values('id', 'name', 'task_count').order_by('-task_count')

        wb = Workbook()
        ws = wb.active
        ws.title = "Activities by Responsible"
        ws.append(["ID", "Name", "Task Count"])

        for user_task in responsible_tasks:
            ws.append([user_task['id'], user_task['name'], user_task['task_count']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=activities_by_responsible.xlsx'
        wb.save(response)

        return response

    def export_to_json(self):
        # Quantidade de atividades por responsável
        responsible_tasks = User.objects.annotate(
            task_count=Count('tasks_responsible')
        ).values('id', 'name', 'task_count').order_by('-task_count')


        data = list(responsible_tasks)

        return Response(data)
    
class LateTasksView(APIView):

    def get(self, request):
        current_date = timezone.now().date()
        format = request.GET.get('formato')

        # Consulta para obter o número de tarefas atrasadas para cada usuário
        late_tasks = User.objects.annotate(
            late_count=Count('tasks_responsible', filter=Q(tasks_responsible__deadline__lt=current_date, tasks_responsible__completed=False)),
            finished_late_count=Count('tasks_responsible', filter=Q(tasks_responsible__deadline__lt=F('tasks_responsible__finished_in')))
        ).values('id', 'name', 'late_count', 'finished_late_count')

        if format == 'excel':
            return self.export_to_excel(late_tasks)
        elif format == 'json':
            return self.export_to_json(late_tasks)
        else:
            return Response({"detail": "Formato não suportado."}, status=status.HTTP_400_BAD_REQUEST)

    def export_to_excel(self, late_tasks):

        wb = Workbook()
        ws = wb.active
        ws.title = "Late Tasks Report"
        ws.append(["ID", "Name", "Late Count", "Finished Late Count"])

        for task in late_tasks:
            ws.append([task['id'], task['name'], task['late_count'], task['finished_late_count']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=late_tasks_report.xlsx'
        wb.save(response)

        return response

    def export_to_json(self, late_tasks):

        data = list(late_tasks)

        return Response(data)

class UserFinishedOwnTasksView(APIView):

    def get(self, request):
        format = request.GET.get('formato')
        # Obter o número de tarefas que o usuário foi responsável e finalizou
        user_finished_own_tasks = User.objects.annotate(
            own_finished_count=Count('tasks_responsible', filter=Q(tasks_responsible__finished_by=F('pk')))
        ).values('id', 'name', 'own_finished_count')

        if format == 'excel':
            return self.export_to_excel(user_finished_own_tasks)
        elif format == 'json':
            return self.export_to_json(user_finished_own_tasks)
        else:
            return Response({"detail": "Formato não suportado."}, status=status.HTTP_400_BAD_REQUEST)

    def export_to_excel(self, user_finished_own_tasks):
 
        wb = Workbook()
        ws = wb.active
        ws.title = "User Finished Own Tasks Report"
        ws.append(["ID", "Name", "Own Finished Count"])

        for user_task in user_finished_own_tasks:
            ws.append([user_task['id'], user_task['name'], user_task['own_finished_count']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=user_finished_own_tasks_report.xlsx'
        wb.save(response)

        return response

    def export_to_json(self, user_finished_own_tasks):
 
        data = list(user_finished_own_tasks)

        return Response(data)
    
class UserCreatedAndFinishedTasksView(APIView):

    def get(self, request):
        format = request.GET.get('formato')
        # Quantas tarefas que o usuário estava como criador, foi ele quem finalizou a tarefa.
        user_created_and_finished_tasks = User.objects.annotate(
            created_and_finished_count=Count('task_profiles_created', filter=Q(task_profiles_created__finished_by=F('pk')))
        ).values('id', 'name', 'created_and_finished_count')
        
        if format == 'excel':
            return self.export_to_excel(user_created_and_finished_tasks)
        elif format == 'json':
            return self.export_to_json(user_created_and_finished_tasks)
        else:
            Response({"Detail": "Formato não suportado"}, status=status.HTTP_400_BAD_REQUEST)
            
    def export_to_excel(self, user_created_and_finished_tasks):
 
        wb = Workbook()
        ws = wb.active
        ws.title = "User Created and Finished Tasks Report"
        ws.append(["ID", "Name", "Created and Finished Count"])

        for user_task in user_created_and_finished_tasks:
            ws.append([user_task['id'], user_task['name'], user_task['created_and_finished_count']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=user_created_and_finished_tasks_report.xlsx'
        wb.save(response)

        return response

    def export_to_json(self, user_created_and_finished_tasks):

        data = list(user_created_and_finished_tasks)

        return Response(data)

class TaskViewsSet(viewsets.ModelViewSet):
    serializer_class = TasksSerializer
    queryset = task.TaskProfile.objects.all()
    filter_backends = [DjangoFilterBackend] 
    filterset_fields = ['created_in', 'finished_in']
    
    
# ---------------------------------------------------------------------------------------------------------------------------------

# class ExcelExportView:
#     def create_excel(self, data, headers, filename):
#         # Criar um arquivo Excel em memória
#         wb = Workbook()
#         ws = wb.active
#         ws.title = filename

#         # Adicionar os cabeçalhos
#         ws.append(headers)

#         # Adicionar os dados
#         for item in data:
#             ws.append(list(item.values()))

#         # Criar a resposta HTTP com o conteúdo do arquivo Excel
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

#         # Salvar o workbook diretamente na resposta HTTP
#         wb.save(response)

#         return response